import pandas as pd
from openpyxl import load_workbook
from fpdf import FPDF

# Load the Excel file
file_path = 'path_to_your_excel_file.xlsx'
df = pd.read_excel(file_path)

# Remove columns
columns_to_remove = ['Column1', 'Column2', 'Column3']  # Replace with the names of the columns you want to remove
df = df.drop(columns=columns_to_remove)

# Increase column size
column_to_resize = 'Column4'  # Replace with the name of the column you want to resize
new_column_width = 20  # Replace with the desired width for the column
ws = load_workbook(file_path).active
ws.column_dimensions[column_to_resize].width = new_column_width
ws.column_dimensions['B'].width = new_column_width  # Example: Resizing column 'B'

# Save modified Excel file
df.to_excel('modified_excel_file.xlsx', index=False)

# Convert Excel to PDF
pdf = FPDF()
xls_file = pd.ExcelFile('modified_excel_file.xlsx')
for sheet_name in xls_file.sheet_names:
    df = xls_file.parse(sheet_name)
    pdf.add_page()
    pdf.set_font("Arial", size=10)
    for row in df.iterrows():
        pdf.cell(0, 10, txt=row[1].to_string(), ln=True)
pdf.output('output_file.pdf')
